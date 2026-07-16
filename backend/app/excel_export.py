from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font

from .models import Car, Event
from .schemas import StatsResponse


XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EVENT_TYPE_LABELS = {
    "fuel": "Заправка",
    "repair": "Ремонт",
    "trip": "Поездка",
    "issue": "Проблема",
}
PERIOD_LABELS = ("Неделя", "Месяц", "Всё время")


def build_export_filename(car: Car, now: datetime) -> str:
    return f"LAMBA_{_safe_filename_part(car.brand)}_{_safe_filename_part(car.model)}_{now:%Y-%m-%d}.xlsx"


def build_workbook(car: Car, events: list[Event], stats: StatsResponse) -> BytesIO:
    workbook = Workbook()
    vehicle_sheet = workbook.active
    vehicle_sheet.title = "Автомобиль"
    _append_vehicle_sheet(vehicle_sheet, car)
    _append_history_sheet(workbook.create_sheet("История событий"), events)
    _append_statistics_sheet(workbook.create_sheet("Статистика"), stats)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _append_vehicle_sheet(sheet, car: Car) -> None:
    rows = (
        ("Марка", car.brand),
        ("Модель", car.model),
        ("Год выпуска", car.production_year),
        ("Текущий пробег", car.current_mileage),
        ("Дата создания профиля", car.created_at.strftime("%d.%m.%Y")),
    )
    for row in rows:
        sheet.append(row)
    _style_header_column(sheet)
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 32


def _append_history_sheet(sheet, events: list[Event]) -> None:
    headers = (
        "Дата",
        "Тип события",
        "Описание",
        "Сумма, ₽",
        "Топливо, л",
        "Пробег, км",
        "Начальный одометр, км",
        "Конечный одометр, км",
        "Расстояние поездки, км",
        "Наличие фото",
    )
    sheet.append(headers)
    for event in events:
        sheet.append(
            (
                event.created_at.strftime("%d.%m.%Y"),
                EVENT_TYPE_LABELS[event.type],
                event.description,
                _number(event.amount),
                _number(event.fuel_liters),
                _number(event.mileage),
                event.odometer_start,
                event.odometer_end,
                _number(event.trip_distance),
                "Да" if event.photo_path else "Нет",
            )
        )
    _style_header_row(sheet)
    for column, width in enumerate((14, 18, 35, 14, 14, 16, 24, 24, 25, 16), start=1):
        sheet.column_dimensions[chr(64 + column)].width = width
    sheet.freeze_panes = "A2"


def _append_statistics_sheet(sheet, stats: StatsResponse) -> None:
    periods = (stats.week, stats.month, stats.all_time)
    sheet.append(("Показатель", *PERIOD_LABELS))
    rows = (
        ("Пробег, км", [period.mileage_km for period in periods]),
        ("Расходы, ₽", [period.expenses_rub for period in periods]),
        ("Расходы на топливо, ₽", [period.fuel_expenses for period in periods]),
        ("Расходы на ремонт, ₽", [period.repair_expenses for period in periods]),
        ("Топливо, л", [period.fuel_liters for period in periods]),
        ("Количество записей", [period.records_count for period in periods]),
        (
            "Средний расход топлива, л/100 км",
            [period.avg_fuel_consumption_l_per_100km for period in periods],
        ),
        (
            "Средний расход, ₽/км",
            [period.avg_expense_consumption for period in periods],
        ),
    )
    for label, values in rows:
        sheet.append((label, *(_number(value) for value in values)))
    _style_header_row(sheet)
    sheet.column_dimensions["A"].width = 38
    for column in ("B", "C", "D"):
        sheet.column_dimensions[column].width = 18
    _add_chart(sheet, 3, "Расходы по периодам", "Расходы, ₽", "E2")
    _add_chart(sheet, 2, "Пробег по периодам", "Пробег, км", "E18")


def _add_chart(sheet, row: int, title: str, axis_title: str, anchor: str) -> None:
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = axis_title
    chart.x_axis.title = "Период"
    chart.add_data(
        Reference(sheet, min_col=2, max_col=4, min_row=row, max_row=row),
        from_rows=True,
        titles_from_data=False,
    )
    chart.set_categories(Reference(sheet, min_col=2, max_col=4, min_row=1, max_row=1))
    chart.series[0].tx = SeriesLabel(v=axis_title)
    sheet.add_chart(chart, anchor)


def _style_header_row(sheet) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True)


def _style_header_column(sheet) -> None:
    for row in sheet.iter_rows(min_col=1, max_col=1):
        row[0].font = Font(bold=True)


def _number(value: Decimal | int | None) -> int | float | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return int(value) if value == value.to_integral_value() else float(value)
    return value


def _safe_filename_part(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", normalized).strip("._-")
    return cleaned or "vehicle"
