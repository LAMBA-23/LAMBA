from io import BytesIO

from openpyxl import load_workbook


def _register_user(client, username: str) -> int:
    response = client.post(
        "/auth/register",
        json={"username": username, "password": "password123"},
    )
    assert response.status_code == 201
    return response.json()["user_id"]


def _create_vehicle(client, user_id: int, brand: str = "Toyota", model: str = "Camry"):
    response = client.post(
        "/vehicle",
        json={
            "user_id": user_id,
            "brand": brand,
            "model": model,
            "production_year": 2020,
            "current_mileage": 100000,
        },
    )
    assert response.status_code == 201


def _create_event(client, user_id: int, **overrides):
    payload = {
        "type": "fuel",
        "description": "Fuel",
        "amount": 1000,
        "fuel_liters": 20,
        "mileage": 0,
    }
    payload.update(overrides)
    response = client.post(f"/events?user_id={user_id}", json=payload)
    assert response.status_code == 200


def _workbook(response):
    return load_workbook(BytesIO(response.content))


def test_export_returns_an_xlsx_attachment_for_the_requested_user(client):
    user_id = _register_user(client, "export-owner")
    _create_vehicle(client, user_id, brand="Toyota", model="Camry")

    response = client.get(f"/data/export.xlsx?user_id={user_id}")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment;" in response.headers["content-disposition"]
    assert 'filename="LAMBA_Toyota_Camry_' in response.headers["content-disposition"]
    assert response.headers["content-disposition"].endswith('.xlsx"')


def test_export_contains_russian_sheets_vehicle_events_statistics_and_charts(client):
    user_id = _register_user(client, "export-content")
    _create_vehicle(client, user_id)
    _create_event(
        client,
        user_id,
        type="fuel",
        description="Full tank",
        amount=2500,
        fuel_liters=40.5,
        mileage=100050,
    )
    _create_event(
        client,
        user_id,
        type="trip",
        description="Weekend trip",
        amount=0,
        mileage=50,
        odometer_start=100050,
        odometer_end=100100,
    )
    _create_event(client, user_id, type="issue", description="Engine light", amount=0)

    workbook = _workbook(client.get(f"/data/export.xlsx?user_id={user_id}"))

    assert workbook.sheetnames == ["Автомобиль", "История событий", "Статистика"]
    vehicle = workbook["Автомобиль"]
    assert [vehicle.cell(row, 1).value for row in range(1, 6)] == [
        "Марка",
        "Модель",
        "Год выпуска",
        "Текущий пробег",
        "Дата создания профиля",
    ]
    assert [vehicle.cell(row, 2).value for row in range(1, 5)] == [
        "Toyota",
        "Camry",
        2020,
        100000,
    ]

    history = workbook["История событий"]
    assert [cell.value for cell in history[1]] == [
        "Дата",
        "Тип события",
        "Описание",
        "Сумма, ₽",
        "Топливо, л",
        "Пробег, км",
        "Начальный одометр, км",
        "Конечный одометр, км",
        "Расстояние поездки, км",
        "Наличие фото",
    ]
    assert [history.cell(row, 2).value for row in range(2, 5)] == [
        "Заправка",
        "Поездка",
        "Проблема",
    ]
    assert [history.cell(row, 10).value for row in range(2, 5)] == ["Нет", "Нет", "Нет"]
    assert all("/events/" not in str(cell.value) for row in history for cell in row)

    stats = workbook["Статистика"]
    assert [cell.value for cell in stats[1]] == [
        "Показатель",
        "Неделя",
        "Месяц",
        "Всё время",
    ]
    assert {stats.cell(row, 1).value for row in range(2, stats.max_row + 1)} >= {
        "Пробег, км",
        "Расходы, ₽",
        "Количество записей",
    }
    assert len(stats._charts) == 2
    assert {chart.title.tx.rich.p[0].r[0].t for chart in stats._charts} == {
        "Расходы по периодам",
        "Пробег по периодам",
    }
    assert all(chart.legend is not None for chart in stats._charts)
    assert {chart.series[0].tx.v for chart in stats._charts} == {
        "Расходы, ₽",
        "Пробег, км",
    }


def test_export_keeps_empty_history_valid_and_excludes_other_users_data(client):
    owner_id = _register_user(client, "export-empty-owner")
    other_id = _register_user(client, "export-other-owner")
    _create_vehicle(client, owner_id, brand="Lada", model="Vesta")
    _create_vehicle(client, other_id, brand="Secret", model="Car")
    _create_event(client, other_id, description="Private record")

    workbook = _workbook(client.get(f"/data/export.xlsx?user_id={owner_id}"))

    history = workbook["История событий"]
    assert history.max_row == 1
    values = [
        cell.value for sheet in workbook.worksheets for row in sheet for cell in row
    ]
    assert "Secret" not in values
    assert "Private record" not in values
    assert all("id" not in str(value).lower() for value in values if value is not None)
